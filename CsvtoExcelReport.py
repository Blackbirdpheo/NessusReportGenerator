import csv
import openpyxl
import pandas as pd
import os

def input_csv():
 
  input_file = input('Enter the path to the input CSV file: ')
  return input_file





def csv_to_excel(input_file, output_file):
  
  with open(input_file, 'r') as csv_file:
    reader = csv.reader(csv_file)
    
    workbook = openpyxl.Workbook()
    sheet = workbook.active
   
    row_index = 1
    for row in reader:
      if filter_row(row):  
        reordered_row = reorder_columns(row)
        sheet.append(reordered_row)
      
       
        for cell in sheet[row_index]:
          sheet.row_dimensions[row_index].height = 70
          cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
          cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),
                                           right=openpyxl.styles.Side(style='thin'),
                                           top=openpyxl.styles.Side(style='thin'),
                                           bottom=openpyxl.styles.Side(style='thin'))
        
        risk_cell = sheet.cell(row=row_index, column=4)
        if risk_cell.value == "Critical":
          risk_cell.fill = openpyxl.styles.PatternFill(fgColor='800080', patternType='solid')
        elif risk_cell.value == "High":
          risk_cell.fill = openpyxl.styles.PatternFill(fgColor='FF0000', patternType='solid')
        elif risk_cell.value == "Medium":
          risk_cell.fill = openpyxl.styles.PatternFill(fgColor='FFA500', patternType='solid')
        if row_index == 1: 
          for cell in sheet[1]:
            cell.fill = openpyxl.styles.PatternFill(fgColor='0000CCFF', patternType='solid')
            sheet.row_dimensions[row_index].height = 30
    

        row_index += 1
    

  
    workbook.save(output_file)
    
    
        


   
    

#False positives and accepted risk

def filter_row(row):
  return row[1] != 'None' and row[1] != 'Low' and row[5] not in ['PHP Unsupported Version Detection', 'PHP <7.3.24 Multiple Vulnerabilities','Microsoft ASP.NET ValidateRequest Filters Bypass', 'Microsoft ASP.NET MS-DOS Device Name DoS', 'Unsupported Web Server Detection', 'Java RMI Agent Insecure Configuration', 'Apache Tomcat AJP Connector Request Injection (Ghostcat)', 'Unencrypted Telnet Server', 'Unsupported linux kernel version detected in banner reporting (PCI-DSS check)', 'SSL Certificate Cannot Be Trusted', 'SSL Self-Signed Certificate', 'SSL Certificate with Wrong Hostname', 'SSL Certificate Signed Using Weak Hashing Algorithm', 'OpenSSH PCI Disputed Vulnerabilities', 'OpenSSH PCI Disputed Vulnerabilities.', 'OpenSSH >= 2.3.0 AllowTcpForwarding Port Bouncing','Kernel vulnerabilities detected in banner reporting (PCI-DSS check)', 'OpenSSH S/KEY Authentication Account Enumeration', 'OPIE w/ OpenSSH Account Enumeration', 'SSL Certificate Expiry', 'CGI Generic Cross-Site Request Forgery Detection (potential)', 'HSTS Missing From HTTPS Server (RFC 6797)', 'Microsoft ASP.NET ValidateRequest Filters Bypass', 'Microsoft ASP.NET MS-DOS Device Name DoS (PCI-DSS check)', 'HTTP TRACE / TRACK Methods Allowed', 'SSH Server CBC Mode Ciphers Enabled', 'Microsoft SQL Server Unsupported Version Detection (remote check)', 'SNMP \'GETBULK\' Reflection DDoS', 'MySQL User-Defined Functions Multiple Vulnerabilities' ]

def reorder_columns(row):
 
  return [row[2], row[5], row[7], row[1], row[6], row[8], row[0], row[9], row[3], row[4]]


  
input_files = [file for file in os.listdir() if file.endswith('.csv')]

for input_file in input_files:


 file_df = pd.read_csv(input_file)
 file_df_first_record = file_df.drop_duplicates(subset=["Host", "Name", "Description", "Risk", "Synopsis", "Solution", "Protocol", "Port"], keep="first")
 input_file_name = os.path.splitext(input_file)[0]
 output_file = f"{input_file_name}.xlsx"

 file_df_first_record.to_csv(output_file, index=False)
 output_file_name = os.path.splitext(output_file)[0]
 output_filefinal = f"{output_file_name}.xlsx"


 csv_to_excel(output_file, output_filefinal)
